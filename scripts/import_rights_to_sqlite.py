import argparse
import hashlib
import json
import os
import re
import sqlite3
import zipfile
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_DB_PATH = PROJECT_ROOT / "backend" / "database" / "权责清单.sqlite"
RIGHTS_TABLES = ("source_files", "raw_rows", "rights_items", "import_errors")


TARGET_FIELDS = [
    "sequence_no",
    "item_name",
    "subitem_name",
    "power_type",
    "basis",
    "exercising_body",
    "undertaking_org",
    "implementation_level_authority",
    "department_duty",
    "responsibility_content",
    "responsibility_basis",
    "accountability_scope",
    "accountability_situation",
    "remark",
    "status",
]

SYNONYMS = {
    "sequence_no": ["序号", "编号", "序列"],
    "item_name": ["事项名称", "权责事项名称", "权力事项名称", "职权名称", "事项"],
    "subitem_name": ["子项名称", "子项", "事项子项", "实施事项"],
    "power_type": ["权力类型", "权力类别", "职权类型", "事项类型", "权责类型"],
    "basis": ["实施依据", "设定依据", "依据", "法律依据", "权力依据"],
    "exercising_body": ["行使主体", "实施主体", "责任主体", "主体", "部门名称"],
    "undertaking_org": ["承办机构", "承办科室", "责任科室", "办理机构", "内设机构"],
    "implementation_level_authority": ["实施层级及权限", "实施层级", "行使层级", "权限", "层级"],
    "department_duty": ["部门职责", "主要职责", "职责"],
    "responsibility_content": ["责任事项内容", "责任事项", "责任内容", "责任"],
    "responsibility_basis": ["责任事项依据", "责任依据", "追责依据"],
    "accountability_scope": ["追责对象范围", "追责对象", "问责对象", "责任对象范围"],
    "accountability_situation": ["追责情形", "问责情形", "责任追究情形"],
    "remark": ["备注", "说明"],
    "status": ["状态", "是否公开", "有效状态"],
}

HEADER_TERMS = [term for values in SYNONYMS.values() for term in values]


def norm(value):
    if value is None:
        return ""
    text = str(value).replace("\u3000", " ").strip()
    return re.sub(r"\s+", " ", text)


def file_hash(path):
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def guess_year(name):
    match = re.search(r"(20\d{2})", name)
    return match.group(1) if match else None


def guess_department(name):
    text = Path(name).stem
    text = re.sub(r"^附件[:：]?", "", text)
    text = re.sub(r"\s*\(\d+\)\s*$", "", text)
    text = re.sub(r"20\d{2}年?", "", text)
    text = re.sub(r"权责清单.*$", "", text)
    text = re.sub(r"公示.*$", "", text)
    text = text.strip(" -_：:")
    return text or Path(name).stem


def read_json_list(path):
    if not path or not Path(path).exists():
        return []
    data = json.loads(Path(path).read_text(encoding="utf-8-sig"))
    if isinstance(data, dict):
        return [data]
    return data or []


def read_conversion_map(path):
    mapping = {}
    for item in read_json_list(path):
        mapping[str(Path(item["source"]).resolve())] = str(Path(item["converted"]).resolve())
    return mapping


def read_conversion_errors(path):
    errors = {}
    for item in read_json_list(path):
        errors[str(Path(item["source"]).resolve())] = item.get("error", "conversion failed")
    return errors


def header_score(row):
    text = "|".join(norm(cell) for cell in row)
    return sum(1 for term in set(HEADER_TERMS) if term in text)


def find_header(rows):
    best_index = None
    best_score = 0
    for index, row in enumerate(rows[:30]):
        score = header_score(row)
        if score > best_score:
            best_index = index
            best_score = score
    return best_index if best_score >= 2 else None


def map_headers(headers):
    result = {}
    cleaned = [norm(header) for header in headers]
    for field, synonyms in SYNONYMS.items():
        for index, header in enumerate(cleaned):
            if header and any(synonym in header for synonym in synonyms):
                result[field] = index
                break
    return result


def significant(row):
    text = "".join(norm(value) for value in row)
    return bool(text) and len(text) > 1


def row_to_dict(headers, row):
    result = {}
    for index, value in enumerate(row):
        clean = norm(value)
        if not clean:
            continue
        key = headers[index] if index < len(headers) and headers[index] else f"列{index + 1}"
        result[key] = clean
    return result


def standardize(row, header_map):
    output = {field: None for field in TARGET_FIELDS}
    for field, index in header_map.items():
        if index < len(row):
            value = norm(row[index])
            output[field] = value or None
    return output


def should_insert_item(std):
    if std.get("item_name") or std.get("power_type") or std.get("basis"):
        joined = "".join(value or "" for value in std.values())
        return not ("事项名称" in joined and "权力类型" in joined)
    return False


def is_ole_workbook(path):
    with open(path, "rb") as handle:
        return handle.read(8) == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


def worksheet_rows(workbook_path):
    if not is_ole_workbook(workbook_path) and zipfile.is_zipfile(workbook_path):
        # Pass a binary stream so openpyxl detects the file content instead of
        # rejecting valid OOXML workbooks that were saved with a .xls suffix.
        with open(workbook_path, "rb") as handle:
            workbook = load_workbook(handle, data_only=True, read_only=False)
        for worksheet in workbook.worksheets:
            rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
            yield worksheet.title, rows
        return

    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError(
            "Reading legacy .xls workbooks requires xlrd. "
            "Install scripts/requirements.txt first."
        ) from exc

    workbook = xlrd.open_workbook(workbook_path, on_demand=True)
    try:
        for worksheet in workbook.sheets():
            rows = []
            for row_index in range(worksheet.nrows):
                values = []
                for cell in worksheet.row(row_index):
                    value = cell.value
                    if cell.ctype == xlrd.XL_CELL_NUMBER and value.is_integer():
                        value = int(value)
                    elif cell.ctype == xlrd.XL_CELL_DATE:
                        value = xlrd.xldate.xldate_as_datetime(
                            value,
                            workbook.datemode,
                        )
                    values.append(value)
                rows.append(values)
            yield worksheet.name, rows
    finally:
        workbook.release_resources()


def resolve_db_path(value=None):
    configured = value or os.getenv("PLATFORM_DB_PATH")
    if not configured:
        return DEFAULT_DB_PATH
    path = Path(configured).expanduser()
    if not path.is_absolute():
        path = PROJECT_ROOT / path
    return path.resolve()


def connect_database(db_path):
    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    conn.executescript(
        """
CREATE TABLE IF NOT EXISTS source_files (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  file_name TEXT NOT NULL,
  source_path TEXT NOT NULL,
  converted_path TEXT,
  extension TEXT,
  file_sha256 TEXT,
  department_guess TEXT,
  year_guess TEXT,
  import_status TEXT NOT NULL,
  error_message TEXT,
  imported_at TEXT NOT NULL
);
CREATE TABLE IF NOT EXISTS raw_rows (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  source_file_id INTEGER NOT NULL,
  sheet_name TEXT,
  row_number INTEGER,
  raw_json TEXT NOT NULL,
  nonempty_text TEXT,
  FOREIGN KEY(source_file_id) REFERENCES source_files(id)
);
CREATE TABLE IF NOT EXISTS rights_items (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  source_file_id INTEGER NOT NULL,
  sheet_name TEXT,
  source_row_number INTEGER,
  department_guess TEXT,
  year_guess TEXT,
  sequence_no TEXT,
  item_name TEXT,
  subitem_name TEXT,
  power_type TEXT,
  basis TEXT,
  exercising_body TEXT,
  undertaking_org TEXT,
  implementation_level_authority TEXT,
  department_duty TEXT,
  responsibility_content TEXT,
  responsibility_basis TEXT,
  accountability_scope TEXT,
  accountability_situation TEXT,
  remark TEXT,
  status TEXT,
  raw_json TEXT,
  FOREIGN KEY(source_file_id) REFERENCES source_files(id)
);
CREATE TABLE IF NOT EXISTS import_errors (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  source_file_id INTEGER,
  file_name TEXT,
  source_path TEXT,
  stage TEXT,
  error_message TEXT,
  FOREIGN KEY(source_file_id) REFERENCES source_files(id)
);
CREATE INDEX IF NOT EXISTS idx_rights_dept ON rights_items(department_guess);
CREATE INDEX IF NOT EXISTS idx_rights_power ON rights_items(power_type);
CREATE INDEX IF NOT EXISTS idx_rights_item ON rights_items(item_name);
"""
    )
    return conn


def clear_rights_data(conn):
    # Delete only the data owned by this importer. Other platform tables in the
    # same SQLite database must survive every rights-list refresh.
    conn.execute("DELETE FROM import_errors")
    conn.execute("DELETE FROM rights_items")
    conn.execute("DELETE FROM raw_rows")
    conn.execute("DELETE FROM source_files")
    placeholders = ",".join("?" for _ in RIGHTS_TABLES)
    conn.execute(
        f"DELETE FROM sqlite_sequence WHERE name IN ({placeholders})",
        RIGHTS_TABLES,
    )


def insert_source(conn, source, converted, status, error):
    now = datetime.now().isoformat(timespec="seconds")
    cursor = conn.execute(
        """
INSERT INTO source_files(
  file_name, source_path, converted_path, extension, file_sha256,
  department_guess, year_guess, import_status, error_message, imported_at
) VALUES(?,?,?,?,?,?,?,?,?,?)
""",
        (
            source.name,
            str(source),
            converted,
            source.suffix.lower(),
            file_hash(source),
            guess_department(source.name),
            guess_year(source.name),
            status,
            error,
            now,
        ),
    )
    return cursor.lastrowid


def import_workbook(conn, source_id, source, input_path):
    raw_count = 0
    item_count = 0
    department = guess_department(source.name)
    year = guess_year(source.name)
    for sheet_name, rows in worksheet_rows(input_path):
        header_index = find_header(rows)
        if header_index is not None:
            headers = [norm(value) or f"列{index + 1}" for index, value in enumerate(rows[header_index])]
            header_map = map_headers(headers)
            data_start = header_index + 1
        else:
            max_cols = max((len(row) for row in rows), default=0)
            headers = [f"列{index + 1}" for index in range(max_cols)]
            header_map = {}
            data_start = 0

        for row_number, row in enumerate(rows, start=1):
            if not significant(row):
                continue
            raw = row_to_dict(headers, row)
            raw_json = json.dumps(raw, ensure_ascii=False)
            nonempty_text = " | ".join(norm(value) for value in row if norm(value))
            conn.execute(
                "INSERT INTO raw_rows(source_file_id,sheet_name,row_number,raw_json,nonempty_text) VALUES(?,?,?,?,?)",
                (source_id, sheet_name, row_number, raw_json, nonempty_text),
            )
            raw_count += 1
            if row_number <= data_start:
                continue
            std = standardize(row, header_map)
            if should_insert_item(std):
                values = [source_id, sheet_name, row_number, department, year]
                values.extend(std[field] for field in TARGET_FIELDS)
                values.append(raw_json)
                conn.execute(
                    """
INSERT INTO rights_items(
  source_file_id, sheet_name, source_row_number, department_guess, year_guess,
  sequence_no, item_name, subitem_name, power_type, basis, exercising_body, undertaking_org,
  implementation_level_authority, department_duty, responsibility_content, responsibility_basis,
  accountability_scope, accountability_situation, remark, status, raw_json
) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
""",
                    values,
                )
                item_count += 1
    return raw_count, item_count


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--src-dir", required=True)
    parser.add_argument(
        "--db-path",
        help=(
            "SQLite database path. Defaults to PLATFORM_DB_PATH or "
            f"{DEFAULT_DB_PATH}"
        ),
    )
    parser.add_argument("--conversion-map")
    parser.add_argument("--conversion-errors")
    parser.add_argument(
        "--expected-files",
        type=int,
        help="Rollback unless the source directory contains exactly this many workbooks.",
    )
    parser.add_argument(
        "--expected-items",
        type=int,
        help="Rollback unless the import produces exactly this many rights items.",
    )
    args = parser.parse_args()

    src_dir = Path(args.src_dir)
    db_path = resolve_db_path(args.db_path)
    conversion_map = read_conversion_map(args.conversion_map)
    conversion_errors = read_conversion_errors(args.conversion_errors)
    files = sorted(
        [path for path in src_dir.iterdir() if path.is_file() and path.suffix.lower() in (".xls", ".xlsx")],
        key=lambda path: path.name,
    )
    if not files:
        raise ValueError(
            f"No .xls or .xlsx source files found in {src_dir}; "
            "the existing rights data was not changed."
        )
    if args.expected_files is not None and len(files) != args.expected_files:
        raise ValueError(
            f"Expected {args.expected_files} source workbooks but found {len(files)} "
            f"in {src_dir}; the existing rights data was not changed."
        )

    conn = connect_database(db_path)
    summary = []
    try:
        conn.execute("BEGIN IMMEDIATE")
        clear_rights_data(conn)
        for source in files:
            source_resolved = str(source.resolve())
            converted = None
            input_path = None
            error = None
            if source_resolved in conversion_map:
                converted = conversion_map[source_resolved]
                input_path = converted
            elif zipfile.is_zipfile(source) or is_ole_workbook(source):
                input_path = str(source)
            else:
                error = conversion_errors.get(
                    source_resolved,
                    "unsupported or damaged Excel workbook",
                )

            if error:
                source_id = insert_source(conn, source, converted, "error", error)
                conn.execute(
                    "INSERT INTO import_errors(source_file_id,file_name,source_path,stage,error_message) VALUES(?,?,?,?,?)",
                    (source_id, source.name, str(source), "convert", error),
                )
                summary.append((source.name, "error", 0, 0, error))
                continue

            source_id = insert_source(conn, source, converted, "processing", None)
            try:
                raw_count, item_count = import_workbook(conn, source_id, source, input_path)
                conn.execute("UPDATE source_files SET import_status='ok', error_message=NULL WHERE id=?", (source_id,))
                summary.append((source.name, "ok", raw_count, item_count, ""))
            except Exception as exc:
                message = str(exc)
                conn.execute("UPDATE source_files SET import_status='error', error_message=? WHERE id=?", (message, source_id))
                conn.execute(
                    "INSERT INTO import_errors(source_file_id,file_name,source_path,stage,error_message) VALUES(?,?,?,?,?)",
                    (source_id, source.name, str(source), "parse", message),
                )
                summary.append((source.name, "error", 0, 0, message))

        failed_files = [row for row in summary if row[1] != "ok"]
        item_count = conn.execute("SELECT count(*) FROM rights_items").fetchone()[0]
        if failed_files:
            failed_names = ", ".join(row[0] for row in failed_files)
            raise RuntimeError(
                "Rights import was rolled back because one or more files failed: "
                f"{failed_names}"
            )
        if item_count == 0:
            raise RuntimeError(
                "Rights import produced zero rights_items and was rolled back."
            )
        if args.expected_items is not None and item_count != args.expected_items:
            raise RuntimeError(
                f"Expected {args.expected_items} rights_items but imported "
                f"{item_count}; the import was rolled back."
            )

        conn.commit()
        print(f"DB={db_path}")
        print(f"files={conn.execute('select count(*) from source_files').fetchone()[0]}")
        ok_files = conn.execute("select count(*) from source_files where import_status='ok'").fetchone()[0]
        error_files = conn.execute("select count(*) from source_files where import_status='error'").fetchone()[0]
        print(f"ok_files={ok_files}")
        print(f"error_files={error_files}")
        print(f"raw_rows={conn.execute('select count(*) from raw_rows').fetchone()[0]}")
        print(f"rights_items={conn.execute('select count(*) from rights_items').fetchone()[0]}")
        print("by_file=")
        for row in summary:
            print("\t".join(str(value) for value in row))
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    main()
